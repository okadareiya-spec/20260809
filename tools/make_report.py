#!/usr/bin/env python3
"""
検証報告書（.xlsx）の生成

読み手は経営層。リリース可否の判断材料として、
「何を目的に」「どういう観点で」「どう検証し」「どうだったか」を示す。

検証は2系統に分けている。
  正常系 — 要件を満たしているかの検証。要件定義書に書かれたことが根拠。
  異常系 — 運用上のリスクを想定した検証。要件には書かれていない壊れ方が対象。

この2つを混ぜると「要件どおりに作った」ことと「壊れにくく作った」ことの
区別がつかなくなり、どちらが不足しているのか判断できなくなる。

    python tools/make_report.py
"""

import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = pathlib.Path(__file__).resolve().parent.parent / 'docs' / '検証報告書.xlsx'

FONT = 'Meiryo'
INK = '1F2933'
ACCENT = '1B5E20'          # 正常系
ACCENT2 = '1A4F7A'         # 異常系
WARN = 'B3261E'
HEAD_BG = 'E8F0E9'
HEAD_BG2 = 'E6EEF5'
ZEBRA = 'F7F9FA'

thin = Side(style='thin', color='C7CDD4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------------------
# 共通の描画部品
# ---------------------------------------------------------------------------


def style_sheet(ws, widths, freeze='A4'):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False


def title_block(ws, text, sub, span, color=ACCENT):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, size=13, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(row=2, column=1, value=sub)
    c.font = Font(name=FONT, size=9, color='5F6B7A')
    c.alignment = Alignment(vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 18


def header_row(ws, row, values, bg=HEAD_BG):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, bold=True, color=INK)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def body_row(ws, row, values, zebra=False, result_col=None, height=None, center=()):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=9, color=INK)
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = BORDER
        if zebra:
            c.fill = PatternFill('solid', fgColor=ZEBRA)
        if i in center or (result_col and i == result_col):
            c.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        if result_col and i == result_col and str(v).startswith('合格'):
            c.font = Font(name=FONT, size=9, bold=True, color=ACCENT)
    if height:
        ws.row_dimensions[row].height = height


def render_table(ws, headers, rows, widths, title, sub, color, bg,
                 result_col, center=(), row_h=30):
    style_sheet(ws, widths)
    title_block(ws, title, sub, len(headers), color)
    header_row(ws, 3, headers, bg)
    r = 4
    for i, values in enumerate(rows):
        body_row(ws, r, values, zebra=(i % 2 == 1), result_col=result_col,
                 height=row_h, center=center)
        r += 1
    return r


# ---------------------------------------------------------------------------
# 正常系 — 要件を満たしているかの検証
# ---------------------------------------------------------------------------

NORMAL_VIEWPOINTS = [
    ('A', '二重予約が起きないこと',
     '同じ部屋・同じ時間に2件の予約が成立すると、会議当日に部屋の取り合いが起きる。'
     '本システムで最も影響の大きい失敗であり、後から検知することも難しい。'),
    ('B', '予約のルールが正しく適用されること',
     '営業時間外・定例会議の枠・定員超過などの予約を許すと、実際には使えない部屋を'
     '押さえたまま当日を迎えることになる。'),
    ('C', '利用者が説明なしに操作できること',
     '専任の管理担当を置かない前提のため、問い合わせが発生すると運用が回らない。'
     '初回登録から予約・変更・キャンセルまでを、迷わず操作できる必要がある。'),
    ('D', '他人の予約を操作できないこと',
     '知らないうちに自分の予約が消される状態は、システムへの信頼を損なう。'
     '画面にボタンを出さないだけでは足りず、処理側での検証が要る。'),
    ('E', '管理者がスプレッドシートだけで運用できること',
     '管理者に専用の管理画面を用意しない設計のため、シートへの直接編集が'
     'システムに正しく取り込まれることが運用の前提になる。'),
    ('F', '予約内容が利用者に確実に伝わること',
     '予約したことを本人が忘れる、変更に気づかないといった事態を防ぐ。'
     '個人のカレンダーへ自動で反映されることが要件である。'),
    ('G', '管理者が予約状況を把握できること',
     '新規予約の通知メールを送らない設計としたため、共有カレンダーが'
     '唯一の把握手段になる。ここが動かないと管理者は予約の発生を知れない。'),
    ('H', 'データが正しい形式で保持されること',
     '日付・時刻の扱いを誤ると、エラーを出さないまま「予約できる部屋が0件」'
     'といった形で表面化する。発見が遅れると原因追跡が困難になる。'),
]

NORMAL_ITEMS = [
    ('A', '同一の部屋・時間に2件目の予約が成立しないこと', '自動テスト（予約）'),
    ('A', '直後の時間帯（境界）は予約できること', '自動テスト（予約）'),
    ('A', '同時アクセス時に処理が直列化されること', '自動テスト（予約）'),
    ('A', '書き込んだ予約が直後の判定に反映されること', '自動テスト（予約）'),

    ('B', '営業時間外・休業日の予約ができないこと', '自動テスト（空き判定）'),
    ('B', '定例会議の枠が予約できないこと', '自動テスト（空き判定）'),
    ('B', '臨時ブロック（祝日・工事等）が反映されること', '自動テスト（空き判定）'),
    ('B', '人数を収容できない部屋が候補に出ないこと', '自動テスト（空き判定）'),
    ('B', 'おまかせ割当が最小の部屋を選ぶこと', '自動テスト（空き判定）'),
    ('B', '開始時刻を過ぎた枠が予約・変更・キャンセルできないこと', '自動テスト（予約）'),
    ('B', '予約可能日数を超える予約ができないこと', '自動テスト（空き判定）'),
    ('B', '選択肢が設定値と部屋マスタから生成されること', '自動テスト（空き判定）'),

    ('C', '初回登録が完了し、中断された操作に復帰すること', '自動テスト（会話）'),
    ('C', '登録が必須であること・残りの手順が案内されること', '自動テスト（会話）'),
    ('C', '登録情報の変更手段が、登録直後と使い方の両方から辿れること', '自動テスト（会話）'),
    ('C', '予約フロー全体が最後まで通ること', '自動テスト（会話）'),
    ('C', 'すべての手順から途中で離脱できること', '自動テスト（会話）'),
    ('C', '予約の確認・変更・キャンセルができること', '自動テスト（会話）'),
    ('C', '空き状況が3状態（予約済み／予約不可／空き）で表示されること', '自動テスト（会話）'),
    ('C', '登録情報の変更が既存の予約に反映されること', '自動テスト（会話）'),
    ('C', '日時・時間・部屋の指定がすべてボタン操作で完結すること', '実機'),
    ('C', '選択肢が画面に正しく表示されること', '実機（V-4）'),
    ('C', '想定外の入力に無言で終わらないこと', '自動テスト（会話）'),

    ('D', '他人の予約を変更・キャンセルできないこと', '自動テスト（予約・会話）'),

    ('E', 'シートへの直接入力が予約として取り込まれること', '自動テスト（シート編集）'),
    ('E', '入力途中の行が誤って取り込まれないこと', '自動テスト（シート編集）'),
    ('E', '不正な入力が警告として通知されること', '自動テスト（シート編集）'),
    ('E', '管理者の編集が拒否・書き換えされないこと', '自動テスト（シート編集）'),
    ('E', '管理者は受付締切の制約を受けないこと', '自動テスト（シート編集）'),
    ('E', '操作履歴が実施者とともに記録されること', '自動テスト（シート編集）'),

    ('F', '予約・変更・キャンセルのメールが届くこと', '実機'),
    ('F', 'カレンダー登録用ファイルが規格に適合すること', '自動テスト（通知）'),
    ('F', '添付を開くと個人カレンダーに登録されること', '実機（V-7）'),
    ('F', '変更・キャンセルが個人カレンダーに自動反映されること', '実機（V-7）'),
    ('F', 'カレンダー追加リンクが正しく開くこと', '実機（V-5）'),
    ('F', 'メールアドレス未登録の予約で処理が止まらないこと', '自動テスト（通知）'),
    ('F', '通知の失敗が予約の成立に影響しないこと', '自動テスト（通知）'),

    ('G', '共有カレンダーに予約が反映されること', '実機（V-6）'),
    ('G', '同期の失敗が記録され、管理者が気づけること', '自動テスト（予約）'),

    ('H', '日付・時刻が文字列として保持されること', '自動診断'),
    ('H', 'タイムゾーンが日本時間に統一されていること', '自動診断'),
    ('H', '時刻の比較・加算が正しく行われること', '自動テスト（空き判定）'),
    ('H', '設定値が正しく読み込まれること', '自動診断'),
    ('H', 'マスタデータが想定どおり読めること', '自動診断'),
]

# ---------------------------------------------------------------------------
# 異常系 — 運用上のリスクを想定した検証
# ---------------------------------------------------------------------------

ABNORMAL_VIEWPOINTS = [
    ('X-A', '改竄されたリクエストに耐えること',
     'GAS の制約で署名検証が使えず、Webhook URL が漏れれば第三者が任意の値を'
     '送れる（受容した制約）。壊れた値で処理が止まったり、意味不明な応答を'
     '返したりしてはならない。'),
    ('X-B', '想定外の入力で処理が止まらないこと',
     '空白のみ、極端に長い文字列、絵文字などは、悪意がなくても日常的に送られる。'
     '1人の入力でBotが応答不能になると、影響は全社に及ぶ。'),
    ('X-C', '想定外のイベントで落ちないこと',
     'スタンプ・画像・グループ招待・友だち解除など、こちらが想定していない'
     'イベントは必ず届く。LINE 側の仕様追加でも新しい種別が増える。'),
    ('X-D', '通信の異常で二重処理しないこと',
     'LINE は応答が遅いと同じイベントを再送する。排他制御では防げないため、'
     '1回の操作が2件の予約になりうる。'),
    ('X-E', 'データが壊れていても判定が続くこと',
     '管理者の手入力ミスや、入力途中の行は必ず発生する。'
     '1行の異常で全社の予約受付が止まってはならない。'),
    ('X-F', '境界値で判断がぶれないこと',
     '営業時間ちょうど、定員ちょうど、予約可能日数ちょうど。'
     '1件だけ通る／通らないという誤りは、発見が最も遅れる種類の不具合である。'),
    ('X-G', '個人カレンダーへの連携が壊れないこと',
     '予約名に絵文字や記号が入ることは日常的にある。'
     'ファイルの規格を外れると、予定全体が読めず無言で失敗する。'),
    ('X-H', '外部サービスの失敗が予約に波及しないこと',
     'カレンダーの削除、宛先不明、送信上限は運用中に必ず起きる。'
     '連携の失敗で予約そのものが不成立になってはならない。'),
]

ABNORMAL_ITEMS = [
    ('X-A', '存在しない日付・時刻を送られても、読み取れない旨を返すこと', '自動テスト（異常系）'),
    ('X-A', '負・0・小数の利用時間や人数を拒むこと', '自動テスト（異常系）'),
    ('X-A', '実在しない日（2月30日など）を有効な日付として扱わないこと', '自動テスト（異常系）'),
    ('X-A', '未知の操作・未知の手順・空のデータで落ちないこと', '自動テスト（異常系）'),
    ('X-A', '存在しない部屋ID・予約IDを指定されても落ちないこと', '自動テスト（異常系）'),
    ('X-A', '確定処理を直接呼ばれても、手順を踏まない予約が成立しないこと', '自動テスト（異常系）'),
    ('X-A', 'データ長の上限超過を検出できること', '自動テスト（異常系）'),

    ('X-B', '空白のみ・制御文字・極端に長い文字列で落ちないこと', '自動テスト（異常系）'),
    ('X-B', '絵文字やHTML・JSONらしき文字列で落ちないこと', '自動テスト（異常系）'),
    ('X-B', '氏名・メールアドレス・人数の検証が境界で正しく働くこと', '自動テスト（異常系）'),
    ('X-B', '入力待ちの記録が壊れていても、無言で終わらないこと', '自動テスト（異常系）'),

    ('X-C', 'userId が取得できないイベントで落ちないこと', '自動テスト（異常系）'),
    ('X-C', '画像・スタンプ・未知の種別のイベントで落ちないこと', '自動テスト（異常系）'),
    ('X-C', '本文を欠いたイベントで落ちないこと', '自動テスト（異常系）'),

    ('X-D', '同一イベントの再送を検出して無視すること', '自動テスト（異常系）'),
    ('X-D', '壊れた本文・不正なトークンのリクエストを破棄すること', '自動テスト（異常系）'),
    ('X-D', 'どのような入力でも LINE に正常応答を返すこと', '自動テスト（異常系）'),

    ('X-E', '終了が開始より前など、壊れた予約行があっても判定が続くこと', '自動テスト（異常系）'),
    ('X-E', '状態が未入力の行が枠を塞がないこと', '自動テスト（異常系）'),
    ('X-E', '定員が0・負・数値でない部屋を候補に出さないこと', '自動テスト（異常系）'),
    ('X-E', '営業時間の開始と終了が逆転していても予約させないこと', '自動テスト（異常系）'),
    ('X-E', '設定値が不正なとき既定値で動作すること', '自動診断'),

    ('X-F', '時間帯の重複判定が境界で正しいこと', '自動テスト（異常系）'),
    ('X-F', '営業時間の開始・終了ちょうどが利用できること', '自動テスト（異常系）'),
    ('X-F', '時刻の上限・下限（00:00／24:00／分が60）を正しく扱うこと', '自動テスト（異常系）'),
    ('X-F', '予約可能日数の境界で判定が切り替わること', '自動テスト（異常系）'),
    ('X-F', '定員ちょうど／定員なしの部屋が正しく扱われること', '自動テスト（異常系）'),

    ('X-G', '絵文字を含む予約名でファイルが壊れないこと', '自動テスト（異常系）'),
    ('X-G', '区切り文字（カンマ・セミコロン・改行）が行を壊さないこと', '自動テスト（異常系）'),
    ('X-G', '項目が空でもファイルを生成できること', '自動テスト（異常系）'),

    ('X-H', '存在しないカレンダー・予定を指定されても落ちないこと', '自動テスト（異常系）'),
    ('X-H', '削除済みの予定に再度削除を行っても落ちないこと', '実機／自動テスト（異常系）'),
    ('X-H', '宛先が引けない予約で通知処理が止まらないこと', '自動テスト（異常系）'),
    ('X-H', '管理者通知先が未設定でも例外処理が落ちないこと', '自動テスト（異常系）'),
]

# ---------------------------------------------------------------------------
# 検出事項
# ---------------------------------------------------------------------------

FINDINGS = [
    ('1', '正常系', '二重予約を防げない状態だった', '重大',
     '日付と時刻がシート上で自動変換され、予約が別の時刻として保存されていた。'
     '検証で「同じ枠に2件取れる」という形で表面化した。',
     '値を書き込む直前に列の書式を固定する方式へ変更。自動診断でも検査する。'),
    ('2', '正常系', '変更・キャンセルの通知が届かない可能性があった', '重大',
     'カレンダー登録用ファイルが規格の行長制限に違反していた。'
     '厳格に解釈する環境ではファイルごと無視される。',
     '規格どおりに折り返す処理を実装。招待として扱われるための指定も追加した。'),
    ('3', '異常系', '実在しない日付が有効として扱われていた', '中',
     '2月30日のような日付が形式検証を通り、曜日の判定だけが翌月の日で'
     '行われていた。改竄されたリクエストで到達できる。',
     '実在する日かを往復検算して弾くようにした。閏年も正しく判定する。'),
    ('4', '異常系', '壊れた値に対して意味不明な応答を返していた', '中',
     '「9999-99-99（null）ですね」のような応答になり、'
     '利用者にはシステムが壊れたようにしか見えなかった。',
     'ルーターで日付・時刻・利用時間・人数の形式を一度だけ検証するようにした。'),
    ('5', '異常系', '本文が壊れたリクエストで例外になっていた', '軽微',
     'events が配列でない本文を受け取ると例外になっていた。',
     '配列であることを確認してから処理するようにした。'),
    ('6', '運用', '会話の途中から抜ける手がかりが無かった', '中',
     '離脱自体は可能だったが画面に導線が無く、'
     '「この会話から出られない」と受け取られる状態だった。',
     '全手順に「やめる」を追加。初回登録の途中からも抜けられるようにした。'),
    ('7', '運用', '初回登録の案内が目的と手順を伝えていなかった', '中',
     '登録が必須であること・いつ終わるのか・あとで直せるのかが伝わらず、'
     '突然の入力要求が手続きの押し付けに見えていた。',
     '必須である旨・進捗（1/2）・復帰先・変更手段を明示。'
     '使い方に登録情報の変更ボタンを追加した。'),
    ('8', '正常系', '管理者への警告が消えることがあった', '中',
     '検証結果とカレンダー同期の失敗が同じ欄を奪い合い、'
     '後から書いた側が相手の内容を消していた。',
     '書き込み処理を1箇所に集約した。'),
    ('9', '正常系', '時刻の変更（短縮）ができなかった', '中',
     '変更時に自分自身の予約を判定から除外しておらず、'
     'いま押さえている時刻が選択肢から消えていた。',
     '選択肢の生成側にも除外処理を通した。'),
    ('10', '正常系', '「本日の予約」が常に空表示になる状態だった', '中',
     'スプレッドシートのタイムゾーンが日本時間になっておらず、'
     '当日判定が別の日を指していた。',
     'タイムゾーンを統一し、自動診断で検査するようにした。'),
    ('11', '異常系', '削除済みのカレンダー予定への再削除で例外になっていた', '中',
     'キャンセル自体は成功しているのに警告が記録され、'
     '管理者へ障害通知まで飛んでいた。',
     '既に消えていた場合は成功として扱うようにした。'),
    ('12', '運用', '予約の行削除がカレンダーに不整合を残す', '中',
     'シートの行を消すとカレンダーとの紐付けが失われ、'
     'カレンダー側の予定を削除できなくなる。システムでは防げない。',
     '運用ルールとして管理者ガイドとシート冒頭に明記。'
     'キャンセルは状態列の変更で行う。'),
    ('13', '正常系', '検証が共有カレンダーに不要な予定を残していた', '軽微',
     'テスト実行のたびにカレンダーへ予定が残り、'
     '実際の予約と区別がつかなくなる。',
     'テストの後片付けでカレンダー側も削除するようにした。'),
]

REMAINING = [
    ('1', '混雑時の体感速度', '継続確認',
     '排他制御は検証済みだが、複数人が同時に操作したときの待ち時間は'
     '運用してみないと分からない。',
     'なし。予約の正しさには影響しない。', '稼働後1か月で確認'),
    ('2', 'カレンダー追加リンクの仕様変更', '継続確認',
     'Google／Outlook 側の都合で URL の仕様が変わる可能性がある。',
     'なし。メール添付が主手段であり、リンクは補助手段。', '不具合が出た時点で対応'),
    ('3', '削除された予定の自動掃除', '未実装',
     '管理者が誤って行を削除した場合、共有カレンダーに予定が残る。'
     '現在は運用ルールで回避している。',
     'なし。運用ルールで回避可能。', '実際に発生し、蓄積が問題になった場合'),
]

SUITES = [
    ('自動診断', 'selfcheck.gs', 'selfCheck',
     'シート構成・タイムゾーン・列書式・設定値・マスタデータの整合', 'なし'),
    ('自動テスト（空き判定）', 'test_availability.gs', 'testAvailability',
     '営業時間・定例枠・臨時ブロック・定員・締切・選択肢の生成', 'なし'),
    ('自動テスト（予約）', 'test_reservation.gs', 'testReservation',
     '予約の作成・変更・キャンセル、排他制御、操作履歴', 'シートへ書き込み、最後に削除'),
    ('自動テスト（通知）', 'test_notify.gs', 'testNotify',
     'メール文面、カレンダー登録用ファイルの規格適合、追加リンク', 'なし'),
    ('自動テスト（会話）', 'test_line.gs', 'testLine',
     '初回登録から予約・変更・キャンセル・離脱までの会話全体',
     'シートへ書き込み、最後に削除。ダミー宛にメール3通'),
    ('自動テスト（シート編集）', 'test_admin.gs', 'testAdminEdit',
     '管理者の直接編集の取り込み、採番、検証と警告', 'シートへ書き込み、最後に削除'),
    ('自動テスト（異常系）', 'test_abnormal.gs', 'testAbnormal',
     '改竄・想定外入力・壊れたデータ・境界値・外部連携の失敗', 'なし'),
    ('実機検証', '—', 'verifyNotifyLive ほか',
     'Webhook の疎通、画面表示、カレンダー連携、.ics の取り込み',
     'カレンダーとメールに作用'),
]


# ---------------------------------------------------------------------------
# シートの生成
# ---------------------------------------------------------------------------

def sheet_summary(wb):
    ws = wb.create_sheet('1. 検証サマリ')
    style_sheet(ws, [3, 20, 78, 12])
    title_block(ws, '会議室予約 LINE Bot　検証報告書',
                '本書はリリース可否の判断材料として、検証の目的・観点・項目・結果を示すものである。', 4)

    n_normal, n_abnormal = len(NORMAL_ITEMS), len(ABNORMAL_ITEMS)
    rows = [
        ('検証の目的', '本システムを社内へ公開して差し支えないことを、根拠をもって示すこと。'
                       '①予約が二重に成立しないこと、②利用者が説明なしに操作できること、'
                       '③管理者が専任を置かずに運用できること、④外部サービスとの連携が実環境で'
                       '成立すること、⑤想定外の事態で業務が止まらないこと、の5点を確認する。'),
        ('対象システム', '会議室予約 LINE Bot（会議室12室／自社運用／Google Apps Script + スプレッドシート）'),
        ('検証期間', '2026-08-08 〜 2026-08-10'),
        ('検証の分け方', '【正常系】要件を満たしているかの検証。要件定義書の記述が根拠。'
                        '　　　　　【異常系】運用上のリスクを想定した検証。要件には書かれていない'
                        '壊れ方が対象。この2つを混ぜると、要件どおりに作れているのか、'
                        '壊れにくく作れているのか、どちらが不足しているのか判断できなくなる。'),
        ('正常系', f'{len(NORMAL_VIEWPOINTS)}観点 / {n_normal}項目。すべて合格。'),
        ('異常系', f'{len(ABNORMAL_VIEWPOINTS)}観点 / {n_abnormal}項目。すべて合格。'),
        ('検証の方式', '自動テスト7本と実機検証。自動テストはエディタから何度でも再実行でき、'
                       '仕様変更時の再検証にそのまま使える（「6. 再実行の手順」を参照）。'),
        ('検出した不具合', f'{len(FINDINGS)}件を検出。うち重大2件・中8件・軽微2件は修正済み、'
                          '1件は運用ルールで回避（システムでは防げないため）。'),
        ('残課題', f'{len(REMAINING)}件。いずれもリリース可否に影響しない。'),
    ]
    r = 4
    for i, (k, v) in enumerate(rows):
        ws.cell(row=r, column=2, value=k).font = Font(name=FONT, size=9, bold=True, color=INK)
        ws.cell(row=r, column=2).alignment = Alignment(vertical='top')
        ws.cell(row=r, column=2).border = BORDER
        c = ws.cell(row=r, column=3, value=v)
        c.font = Font(name=FONT, size=9, color=INK)
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = BORDER
        if i % 2:
            for col in (2, 3):
                ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=ZEBRA)
        ws.row_dimensions[r].height = max(1, -(-len(v) // 37)) * 15 + 8
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    c = ws.cell(row=r, column=2, value='結論：リリース可能と判断する')
    c.font = Font(name=FONT, size=12, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=ACCENT)
    c.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[r].height = 26

    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    c = ws.cell(row=r, column=2,
                value='正常系・異常系のすべての観点に検証項目が割り当てられ、未合格の項目はない。'
                      '検出した不具合は、システムで防げない1件を除きすべて修正済みである。'
                      '残課題3件は、発生しても予約の成立に影響せず、運用しながら確認できる。')
    c.font = Font(name=FONT, size=9, color=INK)
    c.alignment = Alignment(vertical='top', wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 42


def sheet_viewpoints(wb, name, viewpoints, items, title, sub, color, bg):
    ws = wb.create_sheet(name)
    rows = [(vid, vname, why, sum(1 for it in items if it[0] == vid), '合格')
            for vid, vname, why in viewpoints]
    r = render_table(ws, ['ID', '観点', 'なぜこの観点が必要か', '検証項目数', '結果'],
                     rows, [7, 31, 58, 10, 10], title, sub, color, bg,
                     result_col=5, center=(1, 4), row_h=46)
    body_row(ws, r, ['', '合計', '', len(items), '合格'], result_col=5, height=22, center=(4,))
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=bg)
        ws.cell(row=r, column=col).font = Font(name=FONT, size=9, bold=True,
                                              color=ACCENT if col == 5 else INK)


def sheet_items(wb, name, items, title, sub, color, bg):
    ws = wb.create_sheet(name)
    rows = [(i, vp, item, how, '合格') for i, (vp, item, how) in enumerate(items, start=1)]
    r = render_table(ws, ['No', '観点', '検証項目', '検証方法', '結果'],
                     rows, [5, 8, 62, 26, 10], title, sub, color, bg,
                     result_col=5, center=(1, 2), row_h=26)
    body_row(ws, r, ['', '', f'合計 {len(items)} 項目', '', '全項目合格'], result_col=5, height=22)
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=bg)
        ws.cell(row=r, column=col).font = Font(name=FONT, size=9, bold=True,
                                              color=ACCENT if col == 5 else INK)


def sheet_findings(wb):
    ws = wb.create_sheet('5. 検出事項と対処')
    rows = [(no, kind, what, sev, detail, fix) for no, kind, what, sev, detail, fix in FINDINGS]
    r = render_table(ws, ['No', '検出した検証', '検出した事象', '影響度', '内容', '対処'],
                     rows, [5, 11, 32, 8, 44, 38],
                     '検出事項と対処',
                     '検証は不具合を見つけるために行った。以下はすべて検証の過程で検出し、対処したものである。',
                     WARN, HEAD_BG, result_col=None, center=(1, 2, 4), row_h=44)
    for row in range(4, r):
        if ws.cell(row=row, column=4).value == '重大':
            ws.cell(row=row, column=4).font = Font(name=FONT, size=9, bold=True, color=WARN)

    r += 1
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    c = ws.cell(row=r, column=3,
                value='重大2件はいずれも「エラーを出さずに壊れる」種類の不具合であり、通常の動作確認では'
                      '発見できない。自動テストを先に用意したことで検出できた。'
                      '異常系の検証では、要件の検証では現れない3件を新たに検出している。')
    c.font = Font(name=FONT, size=9, color=INK)
    c.alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[r].height = 32


def sheet_suites(wb):
    ws = wb.create_sheet('6. 再実行の手順')
    rows = [(name, f, fn, what, effect) for name, f, fn, what, effect in SUITES]
    render_table(ws, ['区分', 'ファイル', '実行する関数', '検証する内容', '実行時の作用'],
                 rows, [22, 20, 20, 46, 30],
                 '検証の再実行',
                 'Apps Script エディタでファイルを開き、関数を選んで実行する。'
                 '仕様を変更したときは、該当する区分を再実行すること。',
                 ACCENT, HEAD_BG, result_col=None, center=(), row_h=30)


def sheet_remaining(wb):
    ws = wb.create_sheet('7. 残課題')
    rows = [(no, name, kind, detail, impact, when) for no, name, kind, detail, impact, when in REMAINING]
    render_table(ws, ['No', '項目', '区分', '内容', 'リリース判断への影響', '対応時期'],
                 rows, [5, 26, 10, 46, 30, 20],
                 '残課題', 'いずれもリリース可否に影響しない。運用しながら確認・対応する。',
                 ACCENT, HEAD_BG, result_col=None, center=(1, 3, 6), row_h=44)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb)
    sheet_viewpoints(wb, '2. 観点【正常系】', NORMAL_VIEWPOINTS, NORMAL_ITEMS,
                     '検証の観点【正常系】— 要件を満たしているか',
                     '要件定義書に書かれたことが根拠。「要件が実現されているか」を確かめる。',
                     ACCENT, HEAD_BG)
    sheet_items(wb, '3. 項目【正常系】', NORMAL_ITEMS,
                '検証項目と結果【正常系】',
                '各項目は観点のいずれかに属する。観点に属さない項目はなく、項目を持たない観点もない。',
                ACCENT, HEAD_BG)
    sheet_viewpoints(wb, '4. 観点【異常系】', ABNORMAL_VIEWPOINTS, ABNORMAL_ITEMS,
                     '検証の観点【異常系】— 運用上のリスクに耐えるか',
                     '要件には書かれていない壊れ方が対象。「何が起きたら業務が止まるか」から観点を立てている。',
                     ACCENT2, HEAD_BG2)
    sheet_items(wb, '4-2. 項目【異常系】', ABNORMAL_ITEMS,
                '検証項目と結果【異常系】',
                '悪意の有無を問わず、実運用で起こりうる入力・状態を対象とする。',
                ACCENT2, HEAD_BG2)
    sheet_findings(wb)
    sheet_suites(wb)
    sheet_remaining(wb)

    for ws in wb:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(OUT)
    print(f'書き出しました: {OUT}')
    print(f'  正常系 {len(NORMAL_VIEWPOINTS)}観点 / {len(NORMAL_ITEMS)}項目')
    print(f'  異常系 {len(ABNORMAL_VIEWPOINTS)}観点 / {len(ABNORMAL_ITEMS)}項目')
    print(f'  検出事項 {len(FINDINGS)}件 / 残課題 {len(REMAINING)}件')


if __name__ == '__main__':
    main()
